import streamlit as st
import pdfplumber
import re
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Till Report Extractor", layout="wide")
st.title("📊 Pharmacy Till Report → Excel Extractor")
st.markdown("Upload one or more PDF till reports to extract key financial fields into a formatted Excel spreadsheet.")

def extract_fields_from_pdf(file):
    fields = {
        "Date": None,
        "Total Customers Served": None,
        "Total Sales (incl. GST)": None,
        "Total GST Collected": None,
        "Taxable Sales (GST Sales)": None,
        "Non-Taxable Sales (GST Free)": None,
        "Cash Sales": None,
        "EFTPOS Sales": None,
        "Less Debtor Charges": None,
        "Plus Debtor Account Payments": None
    }

    with pdfplumber.open(file) as pdf:
        text = pdf.pages[0].extract_text()

    date_match = re.search(r"(\d{2}/\d{2}/\d{2}) \d{2}:\d{2}Date:", text)
    if date_match:
        fields["Date"] = datetime.strptime(date_match.group(1), "%d/%m/%y").date()

    patterns = {
        "Total Customers Served": r"Total Customers Served\s+(\d+)",
        "Total Sales (incl. GST)": r"Total Till Turnover\s+\$?([\d,]+\.\d{2})",
        "Total GST Collected": r"Total GST Collected\s+\$?([\d,]+\.\d{2})",
        "Taxable Sales (GST Sales)": r"Total GST Sales\s+\d+ \$?([\d,]+\.\d{2})",
        "Non-Taxable Sales (GST Free)": r"Total GST Free Sales\s+\d+ \$?([\d,]+\.\d{2})",
        "Cash Sales": r"Cash\s+\$?([\d,]+\.\d{2})",
        "EFTPOS Sales": r"EFTPOS\s+\d+ \$?([\d,]+\.\d{2})",
        "Less Debtor Charges": r"Less Debtor Charges\s+\d+ \$?([\d,]+\.\d{2})",
        "Plus Debtor Account Payments": r"Plus Debtor Account Payments\s+\d+ \$?([\d,]+\.\d{2})",
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            value = match.group(1).replace(",", "")
            fields[key] = float(value)

    return fields

def style_excel(ws):
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    currency_fields = [
        "Total Sales (incl. GST)", "Total GST Collected",
        "Taxable Sales (GST Sales)", "Non-Taxable Sales (GST Free)",
        "Cash Sales", "EFTPOS Sales", "Less Debtor Charges", "Plus Debtor Account Payments"
    ]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.font = header_font
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")
            else:
                if ws.cell(row=1, column=cell.column).value in currency_fields:
                    cell.number_format = '"$"#,##0.00'

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    ws.auto_filter.ref = ws.dimensions

uploaded_files = st.file_uploader("Upload PDF Reports", type="pdf", accept_multiple_files=True)

if uploaded_files:
    data_rows = []
    for file in uploaded_files:
        data = extract_fields_from_pdf(file)
        if data["Date"]:
            data_rows.append(data)
        else:
            st.warning(f"Could not extract date from: {file.name}")

    data_rows.sort(key=lambda x: x["Date"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Till Reports"

    headers = list(data_rows[0].keys())
    ws.append(headers)
    for row in data_rows:
        ws.append([row[h] for h in headers])

    style_excel(ws)

    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)

    st.success("✅ Report generated!")
    st.download_button("📥 Download Excel Report", data=excel_io, file_name="Till_Report_Summary.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")